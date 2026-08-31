"""
BRD-AI: Excel Checklist解析器
直接从 Layout_Checklist_Rev1.0.xlsx 读取规则
"""
import re
import openpyxl
from typing import Dict, Any, List, Optional


def _parse_value_with_unit(text: str) -> dict:
    if not text or not isinstance(text, str):
        return None
    result = {}

    patterns = [
        (r"(\d+\.?\d*)\s*mm", "mm"),
        (r"(\d+\.?\d*)\s*mil", "mil"),
        (r"(\d+\.?\d*)\s*inch", "inch"),
        (r"(\d+\.?\d*)\s*μm", "um"),
        (r"≥\s*(\d+\.?\d*)\s*mm", "mm"),
        (r"≤\s*(\d+\.?\d*)\s*mm", "mm"),
        (r"(\d+\.?\d*)\s*oz", "oz"),
        (r"(\d+\.?\d*)\s*A", "A"),
        (r"(\d+\.?\d*)\s*Ω", "ohm"),
        (r"(\d+\.?\d*)\s*ohm", "ohm"),
    ]
    for pattern, unit in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            result["min"] = float(m.group(1))
            result["unit"] = unit
            break

    multiple_values = re.findall(r"(\d+\.?\d*)\s*mm", text)
    if len(multiple_values) >= 2:
        result["values"] = [float(v) for v in multiple_values]

    return result


def _extract_category(row_data: dict) -> str:
    items = row_data.get("items", "")
    if "Design Input" in items or "设计输入" in items:
        return "design_input"
    if "Components Check" in items or "器件检查" in items:
        return "components"
    if "Function Check" in items or "功能检查" in items:
        return "function"
    if "Thermal" in items or "热设计" in items:
        return "thermal"
    if "Power" in items and "Check" in items:
        return "power"
    if "Constraint Setting" in items or "约束设置" in items:
        return "constraint"
    if "A/D" in items:
        return "analog"
    if "Clock" in items or "High Speed" in items or "时钟" in items:
        return "high_speed"
    if "EMC" in items or "Reliability" in items:
        return "emc"
    if "Power" in items and "GND" in items:
        return "power_gnd"
    if "Keepout" in items or "禁布" in items:
        return "keepout"
    if "Route from Pad" in items or "出线" in items:
        return "route"
    if "Silkscreen" in items or "丝印" in items:
        return "silkscreen"
    if "Bar Code" in items or "条码" in items:
        return "barcode"
    if "Through Hole" in items or "过孔" in items:
        return "via"
    if "Test Points" in items or "测试点" in items:
        return "testpoint"
    if "DRC" in items:
        return "drc"
    if "Optical Anchor" in items or "光学" in items:
        return "fiducial"
    if "Solder Mask" in items or "Material" in items or "阻焊" in items:
        return "solder_mask"
    if "Validation" in items or "Fixture" in items:
        return "fixture"
    if "Standardization" in items or "标准化" in items:
        return "standardization"
    return "other"


def read_checklist_excel(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_names = [s.lower() for s in wb.sheetnames]
    sheet_name = "check list" if "check list" in sheet_names else wb.sheetnames[-1]
    ws = wb[sheet_name]

    header_row = None
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        cell_val = str(ws.cell(row_idx, 3).value or "")
        if "No." in cell_val or "序号" in cell_val:
            header_row = row_idx
            break
    if header_row is None:
        header_row = 3

    items = []
    current_category = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        no_val = ws.cell(row_idx, 3).value
        if no_val is None:
            continue

        category_text = str(ws.cell(row_idx, 1).value or "")
        if category_text.strip():
            current_category = _extract_category({"items": category_text})

        grade = str(ws.cell(row_idx, 4).value or "").strip()
        item_text = str(ws.cell(row_idx, 5).value or "").strip()
        standard_text = str(ws.cell(row_idx, 6).value or "").strip()
        note_text = str(ws.cell(row_idx, 7).value or "").strip()

        if not item_text or item_text == "None":
            continue

        try:
            item_no = int(no_val)
        except (ValueError, TypeError):
            continue

        items.append({
            "no": item_no,
            "category": current_category,
            "grade": grade,
            "items": item_text,
            "standard": standard_text,
            "note": note_text,
        })

    rules = {
        "spacing": {},
        "high_speed": {},
        "power": {},
        "board_edge": {},
        "test_point": {},
        "silkscreen": {},
        "solder_mask": {},
        "differential_impedance": {},
        "components": {},
        "keepout": {},
        "route": {},
        "via": {},
    }

    for item in items:
        category = item["category"]
        standard = item["standard"]
        item_text = item["items"]
        no = item["no"]

        parsed = _parse_value_with_unit(standard)
        if not parsed:
            parsed = _parse_value_with_unit(item_text)

        if not parsed:
            continue

        if category == "constraint" or "spacing" in item_text.lower():
            key = f"rule_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["spacing"][key] = parsed

        if category == "high_speed":
            key = f"hs_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["high_speed"][key] = parsed

        if category == "power" or category == "power_gnd":
            key = f"power_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["power"][key] = parsed

        if category == "keepout" or category == "board_edge":
            key = f"edge_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["board_edge"][key] = parsed

        if category == "testpoint":
            key = f"tp_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["test_point"][key] = parsed

        if category == "silkscreen":
            key = f"ss_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["silkscreen"][key] = parsed

        if category == "solder_mask":
            key = f"sm_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["solder_mask"][key] = parsed

        if category == "components":
            key = f"comp_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["components"][key] = parsed

        if category == "route":
            key = f"route_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["route"][key] = parsed

        if category == "via":
            key = f"via_{no}"
            parsed["source"] = f"checklist_{no}"
            rules["via"][key] = parsed

    total_items = len(items)
    total_rules = sum(len(v) for v in rules.values() if isinstance(v, dict))

    return {
        "excel_path": excel_path,
        "total_items": total_items,
        "total_rules_extracted": total_rules,
        "items": items,
        "rules": rules,
    }


def merge_checklist_with_defaults(checklist_rules: dict, default_rules: dict) -> dict:
    merged = dict(default_rules)
    extracted = checklist_rules.get("rules", {})

    for category in extracted:
        if category not in merged or not isinstance(merged.get(category), dict):
            merged[category] = {}
        for key, value in extracted[category].items():
            if isinstance(value, dict) and "min" in value:
                merged[category][key] = value

    return merged