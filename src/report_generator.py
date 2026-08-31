"""
BRD-AI: 报告生成器 (Module 6)
生成约束报告、DFM报告、汇总报告 (Excel)
"""
import os
from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel_report(
    constraints: dict,
    signal_classification: dict,
    dfm_result: dict,
    board_info: dict,
    output_path: str,
) -> str:
    wb = openpyxl.Workbook()

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def style_data(ws, row, cols):
        for col in range(1, cols + 1):
            ws.cell(row=row, column=col).border = thin_border

    # ==================== Sheet 1: Summary ====================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    summary_data = [
        ("Board Name", board_info.get("board_name", "N/A")),
        ("File Size", f"{board_info.get('file_size_mb', 0)} MB"),
        ("Total Nets", board_info.get("net_count", 0)),
        ("Total Layers", board_info.get("layer_count", 0)),
        ("Net Classes", signal_classification.get("total_classes", 0)),
        ("Diff Pairs Found", signal_classification.get("total_diff_pairs", 0)),
        ("Physical Constraints", len(constraints.get("physical_constraints", []))),
        ("Spacing Constraints", len(constraints.get("spacing_constraints", []))),
        ("Electrical Constraints", len(constraints.get("electrical_constraints", []))),
        ("Via Constraints", len(constraints.get("via_constraints", []))),
        ("DFM Errors", dfm_result.get("error_count", 0)),
        ("DFM Warnings", dfm_result.get("warning_count", 0)),
        ("DFM Passed", "YES" if dfm_result.get("passed", False) else "NO"),
        ("Capability Level", dfm_result.get("capability_level", "N/A")),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    ws_summary.cell(row=1, column=1, value="BRD-AI Constraint Report").font = Font(size=14, bold=True)
    ws_summary.merge_cells("A1:B1")

    for i, (key, val) in enumerate(summary_data, start=3):
        ws_summary.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=val)

    # ==================== Sheet 2: Net Classes ====================
    ws_nc = wb.create_sheet("Net Classes")
    headers = ["Class", "Net Count", "Protocol", "Diff Pair", "Power", "GND", "Width Rule", "Spacing Rule"]
    for col, h in enumerate(headers, 1):
        ws_nc.cell(row=1, column=col, value=h)
    style_header(ws_nc, 1, len(headers))

    nc_data = signal_classification.get("net_classes", {})
    row = 2
    for cls_name, cls_info in sorted(nc_data.items()):
        ws_nc.cell(row=row, column=1, value=cls_name)
        ws_nc.cell(row=row, column=2, value=cls_info.get("count", 0))
        ws_nc.cell(row=row, column=3, value=cls_info.get("protocol", ""))
        ws_nc.cell(row=row, column=4, value="YES" if cls_info.get("is_diff_pair") else "NO")
        ws_nc.cell(row=row, column=5, value="YES" if cls_info.get("is_power") else "NO")
        ws_nc.cell(row=row, column=6, value="YES" if cls_info.get("is_gnd") else "NO")
        ws_nc.cell(row=row, column=7, value=cls_info.get("width_rule", ""))
        ws_nc.cell(row=row, column=8, value=cls_info.get("spacing_rule", ""))
        style_data(ws_nc, row, len(headers))
        row += 1

    # ==================== Sheet 3: Physical Constraints ====================
    ws_phys = wb.create_sheet("Physical Constraints")
    headers = ["Name", "Net Class", "Min Width", "Preferred Width", "Max Width", "Unit"]
    for col, h in enumerate(headers, 1):
        ws_phys.cell(row=1, column=col, value=h)
    style_header(ws_phys, 1, len(headers))

    row = 2
    for phys in constraints.get("physical_constraints", []):
        ws_phys.cell(row=row, column=1, value=phys.get("name", ""))
        ws_phys.cell(row=row, column=2, value=phys.get("net_class", ""))
        ws_phys.cell(row=row, column=3, value=phys.get("width_min", ""))
        ws_phys.cell(row=row, column=4, value=phys.get("width_preferred", ""))
        ws_phys.cell(row=row, column=5, value=phys.get("width_max", ""))
        ws_phys.cell(row=row, column=6, value=phys.get("unit", "mm"))
        style_data(ws_phys, row, len(headers))
        row += 1

    # ==================== Sheet 4: Spacing Constraints ====================
    ws_spc = wb.create_sheet("Spacing Constraints")
    headers = ["Name", "Net Class", "Line-Line", "Line-Pin", "Line-Via", "Unit"]
    for col, h in enumerate(headers, 1):
        ws_spc.cell(row=1, column=col, value=h)
    style_header(ws_spc, 1, len(headers))

    row = 2
    for spc in constraints.get("spacing_constraints", []):
        ws_spc.cell(row=row, column=1, value=spc.get("name", ""))
        ws_spc.cell(row=row, column=2, value=spc.get("net_class", ""))
        ws_spc.cell(row=row, column=3, value=str(spc.get("line_to_line", "")))
        ws_spc.cell(row=row, column=4, value=str(spc.get("line_to_pin", "")))
        ws_spc.cell(row=row, column=5, value=str(spc.get("line_to_via", "")))
        ws_spc.cell(row=row, column=6, value=spc.get("unit", "mm"))
        style_data(ws_spc, row, len(headers))
        row += 1

    # ==================== Sheet 5: Electrical Constraints ====================
    ws_elec = wb.create_sheet("Electrical Constraints")
    headers = ["Name", "Net Class", "Protocol", "Diff Pair", "Impedance", "Tolerance", "Line Width", "Gap"]
    for col, h in enumerate(headers, 1):
        ws_elec.cell(row=1, column=col, value=h)
    style_header(ws_elec, 1, len(headers))

    row = 2
    for elec in constraints.get("electrical_constraints", []):
        ws_elec.cell(row=row, column=1, value=elec.get("name", ""))
        ws_elec.cell(row=row, column=2, value=elec.get("net_class", ""))
        ws_elec.cell(row=row, column=3, value=elec.get("protocol", ""))
        ws_elec.cell(row=row, column=4, value="YES" if elec.get("is_diff_pair") else "NO")
        ws_elec.cell(row=row, column=5, value=elec.get("impedance_target", ""))
        ws_elec.cell(row=row, column=6, value=f"±{elec.get('impedance_tolerance', '')}%" if elec.get("impedance_tolerance") else "")
        ws_elec.cell(row=row, column=7, value=elec.get("min_line_width", "") if elec.get("is_diff_pair") else "")
        ws_elec.cell(row=row, column=8, value=elec.get("min_gap", "") if elec.get("is_diff_pair") else "")
        style_data(ws_elec, row, len(headers))
        row += 1

    # ==================== Sheet 6: DFM Report ====================
    ws_dfm = wb.create_sheet("DFM Report")
    ws_dfm.column_dimensions["A"].width = 10
    ws_dfm.column_dimensions["B"].width = 60

    ws_dfm.cell(row=1, column=1, value="DFM Check Result").font = Font(size=14, bold=True)
    ws_dfm.cell(row=2, column=1, value=f"Capability: {dfm_result.get('capability_level', '')}")
    ws_dfm.cell(row=3, column=1, value=f"Passed: {'YES' if dfm_result.get('passed') else 'NO'}")

    ws_dfm.cell(row=5, column=1, value="ERRORS").font = Font(bold=True, color="FF0000")
    for i, err in enumerate(dfm_result.get("errors", []), start=6):
        cell = ws_dfm.cell(row=i, column=2, value=err)
        cell.fill = error_fill

    warn_start = 6 + len(dfm_result.get("errors", [])) + 1
    ws_dfm.cell(row=warn_start, column=1, value="WARNINGS").font = Font(bold=True, color="FFA500")
    for i, warn in enumerate(dfm_result.get("warnings", []), start=warn_start + 1):
        cell = ws_dfm.cell(row=i, column=2, value=warn)
        cell.fill = warn_fill

    # ==================== Sheet 7: Checklist Rules ====================
    ws_chk = wb.create_sheet("Checklist Rules")
    headers = ["Category", "Rule", "Value", "Unit", "Source"]
    for col, h in enumerate(headers, 1):
        ws_chk.cell(row=1, column=col, value=h)
    style_header(ws_chk, 1, len(headers))

    row = 2
    checklist = constraints.get("checklist_rules", {})
    for category, rules in checklist.items():
        if isinstance(rules, dict):
            for rule_name, rule_val in rules.items():
                if isinstance(rule_val, dict):
                    ws_chk.cell(row=row, column=1, value=category)
                    ws_chk.cell(row=row, column=2, value=rule_name)
                    ws_chk.cell(row=row, column=3, value=str(rule_val.get("min", rule_val.get("impedance", rule_val.get("value", "")))))
                    ws_chk.cell(row=row, column=4, value=rule_val.get("unit", ""))
                    ws_chk.cell(row=row, column=5, value=rule_val.get("source", ""))
                    style_data(ws_chk, row, len(headers))
                    row += 1

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def generate_text_report(
    constraints: dict,
    signal_classification: dict,
    dfm_result: dict,
    board_info: dict,
    output_path: str,
) -> str:
    lines = []
    lines.append("=" * 60)
    lines.append("  BRD-AI: PCB Constraint Generation Report")
    lines.append("=" * 60)
    lines.append(f"  Board: {board_info.get('board_name', 'N/A')}")
    lines.append(f"  Time:  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 60)
    lines.append("")
    lines.append(f"Total Nets:       {board_info.get('net_count', 0)}")
    lines.append(f"Net Classes:      {signal_classification.get('total_classes', 0)}")
    lines.append(f"Diff Pairs:       {signal_classification.get('total_diff_pairs', 0)}")
    lines.append(f"Physical Rules:   {len(constraints.get('physical_constraints', []))}")
    lines.append(f"Spacing Rules:    {len(constraints.get('spacing_constraints', []))}")
    lines.append(f"Electrical Rules: {len(constraints.get('electrical_constraints', []))}")
    lines.append(f"Via Rules:        {len(constraints.get('via_constraints', []))}")
    lines.append("")
    lines.append(f"DFM Passed:       {'YES' if dfm_result.get('passed') else 'NO'}")
    lines.append(f"DFM Errors:       {dfm_result.get('error_count', 0)}")
    lines.append(f"DFM Warnings:     {dfm_result.get('warning_count', 0)}")
    lines.append("")

    if dfm_result.get("errors"):
        lines.append("--- DFM Errors ---")
        for err in dfm_result["errors"]:
            lines.append(f"  [ERROR] {err}")
        lines.append("")

    if dfm_result.get("warnings"):
        lines.append("--- DFM Warnings ---")
        for warn in dfm_result["warnings"]:
            lines.append(f"  [WARN]  {warn}")
        lines.append("")

    lines.append("--- Net Classes ---")
    nc_data = signal_classification.get("net_classes", {})
    for cls_name, cls_info in sorted(nc_data.items()):
        protocol = cls_info.get("protocol") or ""
        diff = "DIFF" if cls_info.get("is_diff_pair") else ""
        pwr = "PWR" if cls_info.get("is_power") else ""
        gnd = "GND" if cls_info.get("is_gnd") else ""
        flags = " ".join(filter(None, [diff, pwr, gnd]))
        lines.append(f"  {cls_name:20s} {cls_info.get('count', 0):4d} nets  {protocol:10s}  {flags}")

    lines.append("")
    lines.append("=" * 60)
    lines.append("  Report Complete")
    lines.append("=" * 60)

    content = "\n".join(lines)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)
    return output_path